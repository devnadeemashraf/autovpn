"""Automation engine for executing web automation workflows using Puppeteer."""

import asyncio
import time
import random
import string
import re
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any
import pandas as pd
import requests
from pyppeteer import launch

from autovpn.core.config import settings
from autovpn.models.automation import Automation, AutomationStep
from autovpn.models.user_login import UserLogin


class AutomationEngine:
    """Engine for executing web automation workflows using Puppeteer."""

    def __init__(self):
        self.browser = None
        self.page = None
        self.context = {}  # Store variables like generated_prefix, credit_balance, etc.

    async def setup_browser(self, headless: bool = None):
        """Setup Puppeteer browser."""
        if headless is None:
            headless = settings.puppeteer_headless

        print("🚀 Launching Puppeteer browser...")

        try:
            # Minimal, stable browser arguments
            args = [
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
            ]

            self.browser = await launch(
                {
                    "headless": headless,
                    "args": args,
                }
            )

            self.page = await self.browser.newPage()
            await self.page.setViewport({"width": 1920, "height": 1080})

            # Set user agent to avoid detection
            await self.page.setUserAgent(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            )

            # Enable JavaScript
            await self.page.setJavaScriptEnabled(True)

            print("✅ Puppeteer browser initialized successfully")

        except Exception as e:
            print(f"❌ Failed to initialize Puppeteer browser: {str(e)}")
            raise Exception(f"Puppeteer browser initialization failed: {str(e)}")

    async def close_browser(self):
        """Close the Puppeteer browser."""
        if self.browser:
            await self.browser.close()

    async def execute_automation(
        self, automation: Automation, user_login: UserLogin, num_profiles: int
    ) -> dict:
        """Execute an automation workflow."""
        try:
            await self.setup_browser()

            # Initialize context with user data
            self.context = {
                "username": user_login.username,
                "password": user_login.password,
                "num_profiles": num_profiles,
                "automation_name": automation.name,
            }

            # Navigate to base URL
            print(f"🌐 Navigating to: {automation.base_url}")
            await self.page.goto(automation.base_url, {"waitUntil": "networkidle0"})
            print(f"✅ Page loaded successfully")

            # Execute each step
            skip_login_steps = False
            for step in sorted(automation.steps, key=lambda x: x.step_order):
                print(f"📋 Executing step {step.step_order}: {step.description}")

                # Skip login steps if user is already logged in
                if skip_login_steps and step.step_order in [
                    3,
                    4,
                    5,
                    6,
                    7,
                ]:  # Login form steps
                    print(f"⏭️ Skipping step {step.step_order} (user already logged in)")
                    continue

                result = await self.execute_step(step)

                # If session check indicates user is logged in, skip subsequent login steps
                if step.action_type == "check_session_status" and result:
                    skip_login_steps = True
                    print(f"✅ User already logged in, will skip login form steps")

            # Return the result file from context
            result_file = self.context.get("result_file")

            return {
                "success": True,
                "result_file": result_file,
                "message": f"Successfully generated {num_profiles} profiles",
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Automation failed: {str(e)}",
            }
        finally:
            await self.close_browser()

    async def execute_step(self, step: AutomationStep):
        """Execute a single automation step."""
        try:
            # Check if page is still valid
            if not self.page or self.page.isClosed():
                raise Exception("Page is closed or invalid")

            result = None

            if step.action_type == "click":
                await self.click_element(step)
            elif step.action_type == "input":
                await self.input_text(step)
            elif step.action_type == "wait":
                await self.wait_action(step)
            elif step.action_type == "navigate":
                await self.navigate_to(step)
            elif step.action_type == "wait_for_element":
                await self.wait_for_element(step)
            elif step.action_type == "check_element_text":
                await self.check_element_text(step)
            elif step.action_type == "check_url_redirect":
                await self.check_url_redirect(step)
            elif step.action_type == "check_session_status":
                result = await self.check_session_status(step)
            elif step.action_type == "get_element_value":
                await self.get_element_value(step)
            elif step.action_type == "check_credit_balance":
                await self.check_credit_balance(step)
            elif step.action_type == "generate_prefix":
                await self.generate_prefix(step)
            elif step.action_type == "submit_form":
                await self.submit_form(step)
            elif step.action_type == "check_operation_success":
                await self.check_operation_success(step)
            elif step.action_type == "find_generated_entry":
                await self.find_generated_entry(step)
            elif step.action_type == "download_csv":
                await self.download_csv(step)
            elif step.action_type == "convert_to_excel":
                await self.convert_to_excel(step)
            else:
                raise ValueError(f"Unknown action type: {step.action_type}")

            return result

        except Exception as e:
            print(f"❌ Step {step.step_order} failed: {str(e)}")
            raise Exception(f"Step {step.step_order} failed: {str(e)}")

    async def click_element(self, step: AutomationStep):
        """Click on an element."""
        element = await self.find_element(step)
        await element.click()

        # Add extra delay in development mode to see what's happening
        if not settings.puppeteer_headless:
            await asyncio.sleep(1)

        if step.wait_time:
            await asyncio.sleep(step.wait_time)

    async def input_text(self, step: AutomationStep):
        """Input text into an element."""
        element = await self.find_element(step)
        await element.click()  # Focus the element

        # Clear the element properly using page.evaluate
        await self.page.evaluate("el => el.value = ''", element)

        # Handle dynamic values from context
        input_value = step.input_value
        if input_value.startswith("{") and input_value.endswith("}"):
            key = input_value[1:-1]  # Remove braces
            input_value = self.context.get(key, input_value)

        # Convert to string to handle integers and other types
        input_value = str(input_value)

        print(f"📝 Inputting text: {input_value}")

        # Use element.type() with proper error handling
        try:
            await element.type(input_value)
        except Exception as e:
            print(f"⚠️ Element.type() failed, trying alternative method: {str(e)}")
            # Alternative: use page.keyboard.type() after focusing
            await element.focus()
            await self.page.keyboard.type(input_value)

        # Add extra delay in development mode to see what's happening
        if not settings.puppeteer_headless:
            await asyncio.sleep(0.5)

        if step.wait_time:
            await asyncio.sleep(step.wait_time)

    async def wait_action(self, step: AutomationStep):
        """Wait for specified time."""
        await asyncio.sleep(step.wait_time or 1)

    async def navigate_to(self, step: AutomationStep):
        """Navigate to a URL."""
        print(f"🌐 Navigating to: {step.input_value}")
        await self.page.goto(step.input_value, {"waitUntil": "networkidle0"})
        print(f"✅ Navigation completed")

        if step.wait_time:
            await asyncio.sleep(step.wait_time)

    async def wait_for_element(self, step: AutomationStep):
        """Wait for an element to be present."""
        print(f"🔍 Waiting for element: {step.xpath or step.css_selector}")

        # Wait for element with timeout
        try:
            if step.xpath:
                await self.page.waitForXPath(step.xpath, {"timeout": 10000})
            elif step.css_selector:
                await self.page.waitForSelector(step.css_selector, {"timeout": 10000})

            print(f"✅ Element found")
        except Exception as e:
            print(f"❌ Element not found: {str(e)}")
            raise Exception(f"Element not found: {step.xpath or step.css_selector}")

        if step.wait_time:
            await asyncio.sleep(step.wait_time)

    async def check_element_text(self, step: AutomationStep):
        """Check if element text matches expected value."""
        element = await self.find_element(step)
        text = await self.page.evaluate("el => el.textContent.trim()", element)
        expected = step.success_indicator

        if text != expected:
            raise Exception(
                f"Element text '{text}' does not match expected '{expected}'"
            )

        if step.wait_time:
            await asyncio.sleep(step.wait_time)

    async def check_url_redirect(self, step: AutomationStep):
        """Check if page has redirected to expected URL."""
        current_url = self.page.url
        expected_url = step.success_indicator

        print(f"🔍 Checking URL redirect: {current_url}")
        print(f"🔍 Expected URL: {expected_url}")

        if expected_url in current_url:
            print(f"✅ URL redirect successful: {current_url}")
            return True
        else:
            raise Exception(
                f"URL redirect failed. Current: {current_url}, Expected: {expected_url}"
            )

    async def check_session_status(self, step: AutomationStep):
        """Check if user is already logged in by checking current URL."""
        current_url = self.page.url
        print(f"🔍 Checking session status: {current_url}")

        # If we're already on dashboard, user is logged in
        if "dashboard" in current_url:
            print(f"✅ User already logged in, skipping login steps")
            return True
        else:
            print(f"ℹ️ User not logged in, proceeding with login")
            return False

    async def get_element_value(self, step: AutomationStep):
        """Get value from an input element."""
        element = await self.find_element(step)
        value = await self.page.evaluate("el => el.value", element)
        self.context["credit_balance"] = int(value) if value else 0

        if step.wait_time:
            await asyncio.sleep(step.wait_time)

    async def check_credit_balance(self, step: AutomationStep):
        """Check if user has enough credit for requested profiles."""
        credit_balance = self.context.get("credit_balance", 0)
        num_profiles = self.context.get("num_profiles", 0)

        if num_profiles > credit_balance:
            raise Exception(
                f"Insufficient credit. Available: {credit_balance}, Requested: {num_profiles}"
            )

        if step.wait_time:
            await asyncio.sleep(step.wait_time)

    async def generate_prefix(self, step: AutomationStep):
        """Generate random 4-character alphanumeric prefix that doesn't start with 0."""
        import string
        import random

        # Generate random 4-character alphanumeric string (lowercase only)
        # First character: 1-9 (no 0)
        # Remaining characters: 0-9, a-z
        first_char = str(random.randint(1, 9))
        remaining_chars = "".join(
            random.choices(string.ascii_lowercase + string.digits, k=3)
        )
        prefix = first_char + remaining_chars

        self.context["generated_prefix"] = prefix
        print(f"🔤 Generated prefix: {prefix}")

        if step.wait_time:
            await asyncio.sleep(step.wait_time)

    async def submit_form(self, step: AutomationStep):
        """Submit a form by finding and clicking the submit button."""
        # Try to find submit button
        submit_selectors = [
            "button[type='submit']",
            "input[type='submit']",
            "button:contains('Submit')",
            "button:contains('Create')",
            "button:contains('Add')",
        ]

        for selector in submit_selectors:
            try:
                submit_button = await self.page.querySelector(selector)
                if submit_button:
                    await submit_button.click()
                    break
            except:
                continue
        else:
            raise Exception("Could not find submit button")

        if step.wait_time:
            await asyncio.sleep(step.wait_time)

    async def check_operation_success(self, step: AutomationStep):
        """Check if operation was successful."""
        element = await self.find_element(step)
        text = await self.page.evaluate("el => el.textContent.trim()", element)

        if "Error" in text:
            # Get error message
            try:
                error_element = await self.page.querySelector(
                    step.xpath + "/following-sibling::text()"
                )
                if error_element:
                    error_msg = await self.page.evaluate(
                        "el => el.textContent.trim()", error_element
                    )
                    raise Exception(f"Operation failed: {error_msg}")
                else:
                    raise Exception(f"Operation failed: {text}")
            except:
                raise Exception(f"Operation failed: {text}")

        if step.wait_time:
            await asyncio.sleep(step.wait_time)

    async def find_generated_entry(self, step: AutomationStep):
        """Find the generated entry in the log table using direct XPath for top row."""
        generated_prefix = self.context.get("generated_prefix")

        # Use direct XPath for top row (row 1)
        prefix_xpath = "/html/body/div[2]/div[3]/section/div[2]/div/div/div/div/div/div/div[2]/div/table/tbody/tr[1]/td[3]/span"
        date_xpath = "/html/body/div[2]/div[3]/section/div[2]/div/div/div/div/div/div/div[2]/div/table/tbody/tr[1]/td[1]/span"

        try:
            # Wait for elements to be present
            await self.page.waitForXPath(prefix_xpath, {"timeout": 10000})
            await self.page.waitForXPath(date_xpath, {"timeout": 10000})

            # Get elements
            prefix_elements = await self.page.xpath(prefix_xpath)
            date_elements = await self.page.xpath(date_xpath)

            if prefix_elements and date_elements:
                prefix_element = prefix_elements[0]
                date_element = date_elements[0]

                # Get prefix text
                prefix_text = await self.page.evaluate(
                    "el => el.textContent.trim()", prefix_element
                )

                print(f"🔍 Found prefix in top row: {prefix_text}")
                print(f"🔍 Expected prefix: {generated_prefix}")

                # Verify it matches our generated prefix
                if prefix_text == generated_prefix:
                    # Store the date for download
                    date_text = await self.page.evaluate(
                        "el => el.textContent.trim()", date_element
                    )
                    print(f"📅 Found date: {date_text}")

                    # Convert date format from "07-18-2025 13:47:52" to "2025-07-18 13:47:52"
                    date_parts = date_text.split(" ")
                    if len(date_parts) == 2:
                        date_part, time_part = date_parts
                        month, day, year = date_part.split("-")
                        formatted_date = f"{year}-{month}-{day} {time_part}"
                        self.context["download_date"] = formatted_date
                        print(
                            f"✅ Successfully found generated entry with prefix {generated_prefix}"
                        )
                        return
                else:
                    raise Exception(
                        f"Prefix mismatch. Expected: {generated_prefix}, Found: {prefix_text}"
                    )
            else:
                raise Exception("Could not find prefix or date elements in top row")
        except Exception as e:
            raise Exception(f"Failed to find generated entry: {str(e)}")

    async def download_csv(self, step: AutomationStep):
        """Download CSV file by clicking the download button and reading from browser download directory."""
        print("🔍 Looking for download button in the table...")

        # Find the download button in the first row
        download_button_xpath = "/html/body/div[2]/div[3]/section/div[2]/div/div/div/div/div/div/div[2]/div/table/tbody/tr[1]/td[5]/a[1]"

        try:
            # Wait for the download button to be present
            await self.page.waitForXPath(download_button_xpath, {"timeout": 10000})

            # Find the download button
            download_buttons = await self.page.xpath(download_button_xpath)
            if not download_buttons:
                raise Exception("Download button not found")

            download_button = download_buttons[0]

            print("✅ Found download button, clicking...")

            # Set up download handling
            download_dir = Path(settings.download_dir)
            download_dir.mkdir(exist_ok=True)

            # Get the download URL from the button's href
            download_url = await self.page.evaluate("el => el.href", download_button)
            print(f"🔗 Download URL: {download_url}")

            # Click the download button
            await download_button.click()

            # Wait for download to start and complete
            await asyncio.sleep(3)

            # Now we need to find the downloaded file in the browser's download directory
            # Puppeteer downloads to a temporary directory, we need to find it

            # Method 1: Try to get the download path from the browser
            try:
                # Get the browser's download path
                download_path = await self.page.evaluate(
                    """
                    () => {
                        // Try to get the download path from the browser
                        if (window.chrome && window.chrome.downloads) {
                            return window.chrome.downloads.getFileIcon;
                        }
                        return null;
                    }
                """
                )
                print(f"🔍 Browser download path: {download_path}")
            except:
                pass

            # Method 2: Look for the file in common download directories
            import os
            import glob
            from datetime import datetime, timedelta

            # Common download directories to check
            download_dirs = [
                os.path.expanduser("~/Downloads"),  # User downloads
                os.path.expanduser("~/Desktop"),  # Desktop
                os.getcwd() + "/downloads",  # Current working directory
                str(Path(settings.download_dir)),  # Our app download dir
            ]

            print("🔍 Searching for downloaded CSV file...")

            # Look for CSV files modified in the last 30 seconds
            current_time = datetime.now()
            found_files = []

            for search_dir in download_dirs:
                if os.path.exists(search_dir):
                    print(f"🔍 Searching in: {search_dir}")
                    # Look for CSV files
                    csv_pattern = os.path.join(search_dir, "*.csv")
                    csv_files = glob.glob(csv_pattern)

                    for csv_file in csv_files:
                        try:
                            file_time = datetime.fromtimestamp(
                                os.path.getmtime(csv_file)
                            )
                            time_diff = current_time - file_time

                            # Check if file was modified in the last 30 seconds
                            if time_diff.total_seconds() < 30:
                                print(f"✅ Found recent CSV file: {csv_file}")
                                print(f"📅 File modified: {file_time}")
                                print(
                                    f"⏱️ Time difference: {time_diff.total_seconds():.1f} seconds"
                                )
                                found_files.append(
                                    (csv_file, time_diff.total_seconds())
                                )
                        except Exception as e:
                            print(f"⚠️ Error checking file {csv_file}: {e}")

            if found_files:
                # Sort by modification time (most recent first)
                found_files.sort(key=lambda x: x[1])
                downloaded_csv = found_files[0][0]

                print(f"📁 Using most recent CSV file: {downloaded_csv}")

                # Copy the file to our download directory
                import shutil

                timestamp = int(time.time())
                our_csv_file = download_dir / f"vpn_profiles_{timestamp}.csv"

                shutil.copy2(downloaded_csv, our_csv_file)
                print(f"💾 Copied to our directory: {our_csv_file}")

                # Read and validate the CSV content
                with open(our_csv_file, "r", encoding="utf-8") as f:
                    csv_content = f.read()

                print(f"📄 CSV content preview: {csv_content[:200]}...")

                # Validate CSV content - check for CSV headers and structure
                if len(csv_content) > 10 and '"' in csv_content and "," in csv_content:
                    self.context["csv_file"] = str(our_csv_file)
                    print(f"✅ CSV file ready for conversion: {our_csv_file}")
                else:
                    raise Exception("Downloaded file is not valid CSV content")

                # Clean up the original downloaded file
                try:
                    os.remove(downloaded_csv)
                    print(f"🗑️ Cleaned up original file: {downloaded_csv}")
                except Exception as e:
                    print(f"⚠️ Could not clean up original file: {e}")

            else:
                # Fallback: Use the original URL with session cookies
                print("🔄 No downloaded file found, using fallback method...")

                # Get cookies from the current page session
                cookies = await self.page.cookies()
                cookie_dict = {cookie["name"]: cookie["value"] for cookie in cookies}

                # Try to download using requests with session cookies
                response = requests.get(download_url, cookies=cookie_dict)

                print(f"📊 Response status: {response.status_code}")

                if response.status_code == 200 and len(response.content) > 100:
                    # Save CSV file
                    csv_file = download_dir / f"vpn_profiles_{int(time.time())}.csv"
                    with open(csv_file, "wb") as f:
                        f.write(response.content)

                    self.context["csv_file"] = str(csv_file)
                    print(f"💾 CSV file saved (fallback): {csv_file}")
                    print(f"📄 CSV content preview: {response.text[:200]}...")
                else:
                    raise Exception(f"Failed to download CSV: {response.status_code}")

        except Exception as e:
            print(f"❌ Error during download: {str(e)}")
            raise Exception(f"Download failed: {str(e)}")

        if step.wait_time:
            await asyncio.sleep(step.wait_time)

    async def convert_to_excel(self, step: AutomationStep):
        """Convert CSV to Excel format."""
        csv_file = self.context.get("csv_file")
        if not csv_file:
            raise Exception("CSV file not found")

        print(f"📄 Converting CSV to Excel: {csv_file}")

        try:
            # Read CSV with multiple encoding attempts
            df = None
            encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]

            for encoding in encodings:
                try:
                    df = pd.read_csv(
                        csv_file,
                        encoding=encoding,
                        on_bad_lines="skip",
                        engine="python",
                        quotechar='"',
                        skipinitialspace=True,
                    )
                    print(f"✅ Successfully read CSV with {encoding} encoding")
                    break
                except Exception as e:
                    print(f"⚠️ Failed to read with {encoding} encoding: {e}")
                    continue

            if df is None:
                raise Exception("Failed to read CSV file with any encoding")

            print(f"📊 CSV loaded successfully:")
            print(f"   - Rows: {len(df)}")
            print(f"   - Columns: {list(df.columns)}")
            print(f"   - First few rows:")
            print(df.head().to_string())

            # Check if dataframe is empty
            if df.empty:
                raise Exception("CSV file is empty or contains no valid data")

            # Generate filename
            automation_name = self.context.get("automation_name", "AUTOMATION")
            num_profiles = self.context.get("num_profiles", 0)
            current_datetime = datetime.now().strftime("%d%b%Y_%H%M%S").upper()

            excel_filename = (
                f"{automation_name.upper()}_{num_profiles}PIECE_{current_datetime}.xlsx"
            )
            excel_file = Path(settings.download_dir) / excel_filename

            # Convert to Excel with proper formatting and compatibility
            with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
                # Write data to Excel
                df.to_excel(writer, sheet_name="VPN Profiles", index=False)

                # Get the worksheet
                worksheet = writer.sheets["VPN Profiles"]

                # Auto-adjust column widths with better calculation
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    # Check header
                    header_length = len(str(column[0].value)) if column[0].value else 0
                    max_length = max(max_length, header_length)

                    # Check data cells
                    for cell in column[1:]:
                        try:
                            cell_value = (
                                str(cell.value) if cell.value is not None else ""
                            )
                            cell_length = len(cell_value)
                            max_length = max(max_length, cell_length)
                        except:
                            pass

                    # Set column width (minimum 8, maximum 50)
                    adjusted_width = max(8, min(max_length + 2, 50))
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Add some basic formatting
                from openpyxl.styles import Font, PatternFill, Alignment

                # Format header row
                header_font = Font(bold=True)
                header_fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )

                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Center align all data cells
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")

            print(f"✅ Excel file created: {excel_file}")

            # Store result file path
            self.context["result_file"] = str(excel_file)

            # Clean up CSV file
            try:
                Path(csv_file).unlink()
                print(f"🗑️ Cleaned up CSV file: {csv_file}")
            except Exception as e:
                print(f"⚠️ Warning: Could not delete CSV file: {e}")

        except Exception as e:
            print(f"❌ Error converting CSV to Excel: {str(e)}")
            raise Exception(f"CSV to Excel conversion failed: {str(e)}")

        if step.wait_time:
            await asyncio.sleep(step.wait_time)

    async def find_element(self, step: AutomationStep):
        """Find element using xpath or css_selector."""
        try:
            if step.xpath:
                elements = await self.page.xpath(step.xpath)
                if elements:
                    return elements[0]
                else:
                    raise Exception(f"Element not found with xpath: {step.xpath}")
            elif step.css_selector:
                element = await self.page.querySelector(step.css_selector)
                if element:
                    return element
                else:
                    raise Exception(
                        f"Element not found with selector: {step.css_selector}"
                    )
            else:
                raise ValueError("Either xpath or css_selector must be provided")
        except Exception as e:
            print(f"❌ Error finding element: {str(e)}")
            raise e
