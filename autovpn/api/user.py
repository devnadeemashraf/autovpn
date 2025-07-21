"""User API routes."""

from fastapi import APIRouter, Depends, HTTPException, status, Form, Request
from fastapi.responses import HTMLResponse
from sqlmodel import Session, select
from typing import List

from autovpn.core.database import get_session
from autovpn.models.automation import Automation
from autovpn.models.user_login import UserLogin
from autovpn.models.generation_request import GenerationRequest
from autovpn.core.tasks import start_background_task

router = APIRouter()


@router.get("/automations")
async def list_available_automations(session: Session = Depends(get_session)):
    """List available automations for users."""
    statement = select(Automation).where(Automation.is_active == True)
    automations = session.exec(statement).all()

    print(f"Found {len(automations)} automations")
    for automation in automations:
        print(f"- {automation.name} (ID: {automation.id})")

    if not automations:
        return HTMLResponse("<p class='text-gray-500'>No automations available.</p>")

    html = ""
    for automation in automations:
        html += f"""
        <button 
            class="w-full text-left bg-blue-50 hover:bg-blue-100 p-4 rounded mb-2 select-automation"
            data-automation-id="{automation.id}"
            onclick="console.log('Button clicked for automation {automation.id}')"
        >
            <div class="font-semibold">{automation.name}</div>
            <div class="text-sm text-gray-600">{automation.description or 'No description'}</div>
        </button>
        """

    return HTMLResponse(html)


@router.get("/automations/{automation_id}/logins")
async def list_automation_logins(
    automation_id: int, session: Session = Depends(get_session)
):
    """List available logins for a specific automation."""
    statement = select(UserLogin).where(
        UserLogin.automation_id == automation_id, UserLogin.is_active == True
    )
    logins = session.exec(statement).all()

    if not logins:
        return HTMLResponse(
            "<p class='text-gray-500'>No logins available for this automation.</p>"
        )

    html = ""
    for login in logins:
        html += f"""
        <button 
            class="w-full text-left bg-green-50 hover:bg-green-100 p-4 rounded mb-2 select-login"
            data-login-id="{login.id}"
        >
            <div class="font-semibold">{login.display_name or login.username}</div>
            <div class="text-sm text-gray-600">Username: {login.username}</div>
        </button>
        """

    return HTMLResponse(html)


@router.post("/generate")
async def generate_profiles(
    automation_id: int = Form(...),
    user_login_id: int = Form(...),
    num_profiles: int = Form(...),
    session: Session = Depends(get_session),
):
    """Create a new generation request."""
    if num_profiles < 1 or num_profiles > 100:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Number of profiles must be between 1 and 100",
        )

    # Verify automation exists and is active
    automation = session.exec(
        select(Automation).where(
            Automation.id == automation_id, Automation.is_active == True
        )
    ).first()

    if not automation:
        raise HTTPException(status_code=404, detail="Automation not found")

    # Verify user login exists and is active
    user_login = session.exec(
        select(UserLogin).where(
            UserLogin.id == user_login_id,
            UserLogin.automation_id == automation_id,
            UserLogin.is_active == True,
        )
    ).first()

    if not user_login:
        raise HTTPException(status_code=404, detail="User login not found")

    # Create generation request
    request = GenerationRequest(
        automation_id=automation_id,
        user_login_id=user_login_id,
        num_profiles=num_profiles,
        status="pending",
    )
    session.add(request)
    session.commit()
    session.refresh(request)

    # Start background task
    start_background_task(request.id)

    return HTMLResponse(
        f"""
    <div class="bg-blue-100 border border-blue-400 text-blue-700 px-4 py-3 rounded mb-4">
        <div class="flex items-center">
            <div class="animate-spin rounded-full h-4 w-4 border-b-2 border-blue-700 mr-3"></div>
            <div>
                <strong>Generation Started!</strong>
                <p class="text-sm">Request ID: {request.id}</p>
                <p class="text-sm">Status: Pending</p>
            </div>
        </div>
        <div class="mt-3">
            <button 
                onclick="checkStatus({request.id})"
                class="bg-blue-600 text-white px-3 py-1 rounded text-sm hover:bg-blue-700"
            >
                Check Status
            </button>
        </div>
    </div>
    """
    )


@router.get("/requests/{request_id}")
async def get_request_status(request_id: int, session: Session = Depends(get_session)):
    """Get status of a generation request."""
    statement = select(GenerationRequest).where(GenerationRequest.id == request_id)
    request = session.exec(statement).first()

    if not request:
        raise HTTPException(status_code=404, detail="Request not found")

    return request


@router.get("/requests/{request_id}/status")
async def get_request_status_html(
    request_id: int, session: Session = Depends(get_session)
):
    """Get status of a generation request as HTML."""
    statement = select(GenerationRequest).where(GenerationRequest.id == request_id)
    request = session.exec(statement).first()

    if not request:
        return HTMLResponse("<p class='text-red-500'>Request not found</p>")

    status_color = {
        "pending": "blue",
        "running": "yellow",
        "completed": "green",
        "failed": "red",
    }.get(request.status, "gray")

    status_icon = {
        "pending": "⏳",
        "running": "🔄",
        "completed": "✅",
        "failed": "❌",
    }.get(request.status, "❓")

    html = f"""
    <div class="bg-{status_color}-100 border border-{status_color}-400 text-{status_color}-700 px-4 py-3 rounded mb-4">
        <div class="flex items-center">
            <div class="text-2xl mr-3">{status_icon}</div>
            <div>
                <strong>Request #{request.id}</strong>
                <p class="text-sm">Status: {request.status.title()}</p>
                <p class="text-sm">Created: {request.created_at.strftime('%Y-%m-%d %H:%M:%S')}</p>
                {f'<p class="text-sm">Completed: {request.completed_at.strftime("%Y-%m-%d %H:%M:%S")}</p>' if request.completed_at else ''}
                {f'<p class="text-sm">Error: {request.error_message}</p>' if request.error_message else ''}
                {f'<p class="text-sm">Result File: {request.result_file}</p>' if request.result_file else ''}
            </div>
        </div>
        {f'<div class="mt-3"><a href="/download/{request.result_file}" class="bg-green-600 text-white px-3 py-1 rounded text-sm hover:bg-green-700">Download Result</a></div>' if request.result_file else ''}
    </div>
    """

    return HTMLResponse(html)


@router.get("/rate-limit-status")
async def get_rate_limit_status(request: Request):
    """Get current rate limit status for the client IP."""
    from autovpn.core.rate_limiter import rate_limiter, get_client_ip

    client_ip = get_client_ip(request)
    remaining_requests = rate_limiter.get_remaining_requests(client_ip)

    return {
        "ip": client_ip,
        "limit": 5,
        "remaining_requests": remaining_requests,
        "reset_time": "60 seconds from last request",
    }


@router.get("/test-csv-conversion")
async def test_csv_conversion():
    """Test CSV to XLSX conversion with the provided CSV file."""
    import pandas as pd
    from pathlib import Path
    from datetime import datetime
    from openpyxl.styles import Font, PatternFill, Alignment
    from autovpn.core.config import settings

    try:
        # Use the provided CSV file
        csv_file = Path("downloads/vpn_profiles_1753055250.csv")

        if not csv_file.exists():
            raise HTTPException(
                status_code=404, detail=f"CSV file not found: {csv_file}"
            )

        print(f"📄 Testing CSV to Excel conversion: {csv_file}")

        # Read CSV with multiple encoding attempts
        df = None
        encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]

        for encoding in encodings:
            try:
                df = pd.read_csv(
                    csv_file,
                    encoding=encoding,
                    on_bad_lines="skip",
                    engine="python",
                    quotechar='"',
                    skipinitialspace=True,
                )
                print(f"✅ Successfully read CSV with {encoding} encoding")
                break
            except Exception as e:
                print(f"⚠️ Failed to read with {encoding} encoding: {e}")
                continue

        if df is None:
            raise HTTPException(
                status_code=500, detail="Failed to read CSV file with any encoding"
            )

        print(f"📊 CSV loaded successfully:")
        print(f"   - Rows: {len(df)}")
        print(f"   - Columns: {list(df.columns)}")
        print(f"   - First few rows:")
        print(df.head().to_string())

        # Check if dataframe is empty
        if df.empty:
            raise HTTPException(
                status_code=500, detail="CSV file is empty or contains no valid data"
            )

        # Generate test filename
        test_filename = (
            f"TEST_CONVERSION_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        excel_file = Path(settings.download_dir) / test_filename

        # Convert to Excel with proper formatting and compatibility
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            # Write data to Excel
            df.to_excel(writer, sheet_name="VPN Profiles", index=False)

            # Get the worksheet
            worksheet = writer.sheets["VPN Profiles"]

            # Auto-adjust column widths with better calculation
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                # Check header
                header_length = len(str(column[0].value)) if column[0].value else 0
                max_length = max(max_length, header_length)

                # Check data cells
                for cell in column[1:]:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        cell_length = len(cell_value)
                        max_length = max(max_length, cell_length)
                    except:
                        pass

                # Set column width (minimum 8, maximum 50)
                adjusted_width = max(8, min(max_length + 2, 50))
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Add some basic formatting
            # Format header row
            header_font = Font(bold=True)
            header_fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Center align all data cells
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")

        print(f"✅ Excel file created: {excel_file}")

        return {
            "success": True,
            "message": "CSV to Excel conversion test completed successfully",
            "excel_file": str(excel_file),
            "csv_rows": len(df),
            "csv_columns": list(df.columns),
            "excel_file_size": excel_file.stat().st_size if excel_file.exists() else 0,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error in CSV conversion test: {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"CSV conversion test failed: {str(e)}"
        )


@router.get("/download-excel/{filename}")
async def download_excel_file(filename: str):
    """Download a converted Excel file."""
    from fastapi.responses import FileResponse
    from pathlib import Path
    from autovpn.core.config import settings

    try:
        # Security check: only allow .xlsx files
        if not filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

        # Prevent directory traversal
        if ".." in filename or "/" in filename:
            raise HTTPException(status_code=400, detail="Invalid filename")

        excel_file = Path(settings.download_dir) / filename

        if not excel_file.exists():
            raise HTTPException(
                status_code=404, detail=f"Excel file not found: {filename}"
            )

        return FileResponse(
            path=str(excel_file),
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error downloading Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Download failed: {str(e)}")
